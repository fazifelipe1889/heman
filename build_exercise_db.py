"""Construye la base de datos de ejercicios desde la planilla de `ejercicios/`.

Fuente de verdad: `C:\\Users\\felds\\Desktop\\ejercicios\\base datos ejercicios(1).xlsx`
+ imágenes `1.jpg … 252.jpg` (algunas `.jfif`) numeradas por POSICIÓN de fila.

Hace tres cosas:
  1. Copia las imágenes a `public/exercises/{posición}.jpg` (renombra .jfif → .jpg).
  2. Genera `supabase/seed.sql` (catálogo global, created_by IS NULL).
  3. Genera `scripts/exercises_data.json` para el sync de la DB remota.

El mapeo imagen↔ejercicio es por POSICIÓN de fila (1-based), no por la columna ID
(que tiene un typo: la fila 87 dice 109). La imagen 3 no existe → ese ejercicio
queda sin imagen (image_url = NULL, la tarjeta cae al ícono por defecto).

Uso:  py build_exercise_db.py
"""

import json
import os
import shutil
import sys
import io

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

XLSX = r"C:\Users\felds\Desktop\ejercicios\base datos ejercicios(1).xlsx"
SRC_IMG_DIR = r"C:\Users\felds\Desktop\ejercicios"
DST_IMG_DIR = os.path.join(os.path.dirname(__file__), "public", "exercises")
IMG_BASE = "/exercises"  # punto único a tocar si se migra a Supabase Storage

# Normalizaciones de datos (limpieza de la planilla) ────────────────────────────
DIFFICULTY_FIX = {"Principide": "Principiante"}
PRIMARY_FIX = {
    "Gluteos (medio/menor)": "Gluteos (medio)",
    "Flexores y Extensores del Antebrazo": "Antebrazo",
}
VALID_DIFFICULTY = {"Principiante", "Intermedio", "Avanzado"}


def cell(ws, r, c):
    v = ws.cell(r, c).value
    if v is None:
        return ""
    return str(v).strip()


def load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Ejercicios"]
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            continue  # fila sin ID = vacía
        rows.append(
            {
                "name": cell(ws, r, 2),
                "primary_muscle_group": cell(ws, r, 3),
                "secondary_raw": cell(ws, r, 4),
                "equipment": cell(ws, r, 5),
                "movement_type": cell(ws, r, 6),
                "difficulty": cell(ws, r, 7),
                "instructions": cell(ws, r, 8),
            }
        )
    return rows


def copy_images(n):
    """Copia ejercicios/{i}.(jpg|jfif) → public/exercises/{i}.jpg. Devuelve set de posiciones con imagen."""
    os.makedirs(DST_IMG_DIR, exist_ok=True)
    present = set()
    for i in range(1, n + 1):
        src = None
        for ext in ("jpg", "jpeg", "jfif", "png"):
            cand = os.path.join(SRC_IMG_DIR, f"{i}.{ext}")
            if os.path.exists(cand):
                src = cand
                break
        if src is None:
            continue
        shutil.copyfile(src, os.path.join(DST_IMG_DIR, f"{i}.jpg"))
        present.add(i)
    return present


def normalize(rows):
    out = []
    for row in rows:
        diff = DIFFICULTY_FIX.get(row["difficulty"], row["difficulty"])
        if diff not in VALID_DIFFICULTY:
            sys.exit(f"Dificultad inválida en '{row['name']}': {diff!r}")
        primary = PRIMARY_FIX.get(row["primary_muscle_group"], row["primary_muscle_group"])
        secondary = [t.strip() for t in row["secondary_raw"].split(",") if t.strip()]
        out.append(
            {
                "name": row["name"],
                "primary_muscle_group": primary,
                "secondary_muscle_groups": secondary,
                "equipment": row["equipment"],
                "movement_type": row["movement_type"],
                "difficulty": diff,
                "instructions": row["instructions"],
            }
        )
    return out


# SQL helpers ────────────────────────────────────────────────────────────────────
def esc(s):
    return s.replace("'", "''")


def arr(items):
    if not items:
        return "ARRAY[]::text[]"
    return "ARRAY[" + ", ".join(f"'{esc(x)}'" for x in items) + "]"


def img_sql(i, present):
    return f"'{IMG_BASE}/{i}.jpg'" if i in present else "NULL"


def main():
    rows = load_rows()
    records = normalize(rows)
    n = len(records)
    present = copy_images(n)

    # adjuntar image_url por posición
    for i, rec in enumerate(records, start=1):
        rec["image_url"] = f"{IMG_BASE}/{i}.jpg" if i in present else None

    # ── seed.sql ────────────────────────────────────────────────────────────────
    lines = [
        "-- seed.sql — Catálogo de ejercicios globales (created_by IS NULL)",
        "-- Generado por build_exercise_db.py desde 'base datos ejercicios(1).xlsx'.",
        "-- NO editar a mano: re-generar con `py build_exercise_db.py`.",
        "",
        "insert into public.exercises",
        "  (name, primary_muscle_group, secondary_muscle_groups, equipment, movement_type, difficulty, instructions, image_url)",
        "values",
    ]
    sql_rows = []
    for i, rec in enumerate(records, start=1):
        sql_rows.append(
            "  ('{name}', '{prim}', {sec}, '{eq}', '{mov}', '{dif}', '{instr}', {img})".format(
                name=esc(rec["name"]),
                prim=esc(rec["primary_muscle_group"]),
                sec=arr(rec["secondary_muscle_groups"]),
                eq=esc(rec["equipment"]),
                mov=esc(rec["movement_type"]),
                dif=esc(rec["difficulty"]),
                instr=esc(rec["instructions"]),
                img=img_sql(i, present),
            )
        )
    lines.append(",\n".join(sql_rows) + ";")
    seed_path = os.path.join(os.path.dirname(__file__), "supabase", "seed.sql")
    with open(seed_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    # ── JSON para el sync remoto ─────────────────────────────────────────────────
    json_path = os.path.join(os.path.dirname(__file__), "scripts", "exercises_data.json")
    os.makedirs(os.path.dirname(json_path), exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    missing = [i for i in range(1, n + 1) if i not in present]
    print(f"✓ {n} ejercicios")
    print(f"✓ imágenes copiadas: {len(present)}  (sin imagen: {missing})")
    print(f"✓ supabase/seed.sql")
    print(f"✓ scripts/exercises_data.json")


if __name__ == "__main__":
    main()
