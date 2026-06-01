from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Ejercicios"

headers = ["ID","Nombre","Grupo Muscular Principal","Musculos Secundarios","Equipamiento","Tipo de Movimiento","Dificultad","Instrucciones Breves"]

exercises = [
# ══════════════════════════════════════════════════════════════════════════════
# PECHO
# ══════════════════════════════════════════════════════════════════════════════
# Barra
["Press de Banca Plano","Pecho","Triceps, Hombros (anterior)","Barra","Empuje horizontal","Intermedio","Agarre a la anchura de los hombros, bajar la barra al pecho y empujar."],
["Press de Banca Inclinado","Pecho","Hombros (anterior), Triceps","Barra","Empuje inclinado","Intermedio","Banco a 30-45, enfatiza la parte superior del pecho."],
["Press de Banca Declinado","Pecho","Triceps, Hombros","Barra","Empuje declinado","Intermedio","Banco declinado, enfatiza la parte inferior del pecho."],
["Press de Banca Agarre Estrecho","Triceps","Pecho, Hombros","Barra","Empuje estrecho","Intermedio","Agarre a la anchura de hombros, codos pegados al cuerpo."],
["Press Guillotina (Neck Press)","Pecho","Triceps","Barra","Empuje horizontal alto","Avanzado","Barra baja hacia el cuello/garganta, maximo estiramiento pectoral superior."],
["Floor Press con Barra","Pecho","Triceps, Hombros","Barra","Empuje horizontal","Intermedio","Acostado en el suelo, limita el rango eliminando el estiramiento de hombro."],
["Press de Banca con Pausa","Pecho","Triceps, Hombros","Barra","Empuje horizontal","Avanzado","Igual que el press de banca pero pausando 1-2 seg en el pecho antes de subir."],
["Press de Banca con Cadenas","Pecho","Triceps, Hombros","Barra","Empuje horizontal con resistencia variable","Avanzado","Cadenas colgadas de la barra para aumentar la resistencia al tope del movimiento."],
["Press de Banca con Bandas","Pecho","Triceps, Hombros","Barra","Empuje horizontal con resistencia variable","Avanzado","Bandas ancladas al suelo añaden resistencia acomodante."],
["Svend Press","Pecho","Hombros (anterior)","Peso Corporal","Aislamiento de aduccion","Principiante","Plato presionado entre las manos frente al pecho, empujar hacia adelante manteniendo la presion."],
["Bradford Press","Hombros","Triceps, Pecho","Barra","Empuje vertical alternado","Avanzado","Pasar la barra de delante de la cabeza a detras sin bloquear, de forma continua."],
# Mancuernas
["Press de Banca con Mancuernas","Pecho","Triceps, Hombros (anterior)","Mancuernas","Empuje horizontal","Principiante","Mayor rango de movimiento que con barra."],
["Press Inclinado con Mancuernas","Pecho","Hombros (anterior), Triceps","Mancuernas","Empuje inclinado","Principiante","Banco a 30-45."],
["Press Declinado con Mancuernas","Pecho","Triceps","Mancuernas","Empuje declinado","Principiante","Banco declinado, codos a 75 grados."],
["Aperturas con Mancuernas Plano","Pecho","Hombros (anterior)","Mancuernas","Aislamiento horizontal","Principiante","Brazos ligeramente flexionados, arco amplio."],
["Aperturas Inclinadas con Mancuernas","Pecho","Hombros (anterior)","Mancuernas","Aislamiento inclinado","Principiante","Banco a 30, arco controlado."],
["Aperturas Declinadas con Mancuernas","Pecho","Hombros (anterior)","Mancuernas","Aislamiento declinado","Principiante","Banco declinado, enfasis en pecho inferior."],
["Pullover con Mancuerna","Pecho","Dorsal, Triceps","Mancuernas","Aislamiento","Intermedio","Tumbado en banco, bajar la mancuerna por detras de la cabeza."],
["Press de Banca con Mancuernas Agarre Neutro","Pecho","Triceps","Mancuernas","Empuje horizontal neutro","Principiante","Palmas enfrentadas, menos estres en hombros."],
["Press de Banca Una Mancuerna (Unilateral)","Pecho","Triceps, Core","Mancuernas","Empuje horizontal unilateral","Intermedio","Un brazo a la vez, mayor demanda de estabilizacion."],
# Maquina
["Press en Maquina Pecho Plano","Pecho","Triceps, Hombros","Maquina","Empuje horizontal","Principiante","Guiado, ideal para principiantes o fatiga."],
["Press en Maquina Pecho Inclinado","Pecho","Hombros (anterior), Triceps","Maquina","Empuje inclinado","Principiante","Angulo fijo, enfasis pecho superior."],
["Pec Deck (Aperturas en Maquina)","Pecho","Hombros (anterior)","Maquina","Aislamiento horizontal","Principiante","Mantener tension constante, ideal para finalizacion."],
["Hammer Strength Press Plano","Pecho","Triceps, Hombros","Maquina","Empuje horizontal articulado","Intermedio","Maquina de palanca, permite movimiento mas natural que guiado."],
["Hammer Strength Press Inclinado","Pecho","Hombros (anterior), Triceps","Maquina","Empuje inclinado articulado","Intermedio","Palanca inclinada, pecho superior."],
["Hammer Strength Press Declinado","Pecho","Triceps","Maquina","Empuje declinado articulado","Intermedio","Palanca declinada, pecho inferior."],
["Press en Maquina Smith Plano","Pecho","Triceps, Hombros","Maquina Smith","Empuje horizontal guiado","Principiante","Movimiento guiado en plano horizontal."],
["Press en Maquina Smith Inclinado","Pecho","Triceps, Hombros","Maquina Smith","Empuje inclinado guiado","Principiante","Movimiento guiado en plano inclinado."],
["Press en Maquina Smith Declinado","Pecho","Triceps","Maquina Smith","Empuje declinado guiado","Principiante","Movimiento guiado en plano declinado."],
["Maquina Convergente de Pecho","Pecho","Triceps, Hombros","Maquina","Empuje convergente","Principiante","Los brazos se juntan al frente imitando el patron de apertura-press."],
# Cable
["Cruce de Poleas Alto a Bajo","Pecho","Hombros (anterior)","Cable","Aislamiento cruzado","Principiante","Poleas altas, cruzar hacia abajo y adentro."],
["Cruce de Poleas Bajo a Alto","Pecho","Hombros (anterior)","Cable","Aislamiento cruzado","Principiante","Poleas bajas, cruzar hacia arriba."],
["Cruce de Poleas Neutro","Pecho","Hombros (anterior)","Cable","Aislamiento cruzado","Principiante","Poleas a altura media."],
["Press con Cable Plano","Pecho","Triceps, Hombros","Cable","Empuje horizontal","Intermedio","Tension constante a diferencia de la barra."],
["Press con Cable Inclinado","Pecho","Hombros (anterior), Triceps","Cable","Empuje inclinado","Intermedio","Polea baja, enfasis pecho superior."],
["Apertura Unilateral en Cable","Pecho","Hombros (anterior)","Cable","Aislamiento unilateral","Intermedio","Un brazo a la vez, mayor rango y contraccion."],
# Peso Corporal
["Flexiones (Push-ups)","Pecho","Triceps, Hombros, Core","Peso Corporal","Empuje horizontal","Principiante","Manos a anchura de hombros, cuerpo recto."],
["Flexiones Inclinadas (Pies Elevados)","Pecho","Pecho Superior, Triceps","Peso Corporal","Empuje declinado","Intermedio","Pies elevados, mayor enfasis pecho superior."],
["Flexiones Declinadas (Manos Elevadas)","Pecho","Pecho Inferior, Triceps","Peso Corporal","Empuje inclinado","Principiante","Manos sobre superficie elevada, pecho inferior."],
["Flexiones con Pausa","Pecho","Triceps, Hombros","Peso Corporal","Empuje horizontal","Intermedio","Pausa de 2 seg al bajar, elimina el reboote elastico."],
["Flexiones Explosivas (Clap Push-ups)","Pecho","Triceps, Hombros","Peso Corporal","Empuje horizontal explosivo","Avanzado","Empujar con fuerza para despegar las manos y aplaudir."],
["Fondos en Paralelas (Pecho)","Pecho","Triceps, Hombros","Peso Corporal","Empuje vertical declinado","Intermedio","Inclinar el torso hacia adelante para mayor enfasis en pecho."],
["Flexiones Arqueras (Archer Push-ups)","Pecho","Triceps","Peso Corporal","Empuje unilateral","Avanzado","Extender un brazo lateralmente mientras el otro hace la flexion."],
["Flexiones con Manos Juntas","Triceps","Pecho","Peso Corporal","Empuje estrecho","Intermedio","Mayor activacion del triceps lateral."],
["Push-up en Anillas","Pecho","Triceps, Core, Hombros","Peso Corporal","Empuje horizontal inestable","Avanzado","Manos en anillas, requiere gran estabilizacion escapular."],

# ══════════════════════════════════════════════════════════════════════════════
# ESPALDA
# ══════════════════════════════════════════════════════════════════════════════
# Barra - pesos muertos
["Peso Muerto Convencional","Espalda Baja","Gluteos, Isquiotibiales, Trapecios, Cuadriceps","Barra","Bisagra de cadera","Avanzado","Barra sobre el mediepie, espalda neutra, empujar el suelo al tirar."],
["Peso Muerto Sumo","Gluteos","Isquiotibiales, Aductores, Cuadriceps, Espalda Baja","Barra","Bisagra de cadera amplia","Avanzado","Pies muy abiertos, puntas hacia afuera, agarre interior, torso mas erguido."],
["Peso Muerto Rumano","Isquiotibiales","Gluteos, Espalda Baja","Barra","Bisagra de cadera","Intermedio","Rodillas ligeramente flexionadas, barra baja por las piernas."],
["Peso Muerto con Barra Trap (Hexagonal)","Espalda Baja","Gluteos, Isquiotibiales, Cuadriceps","Barra Hexagonal","Bisagra de cadera","Intermedio","Barra hexagonal, agarre neutro, postura intermedia entre convencional y sentadilla."],
["Peso Muerto Piernas Rigidas (SLDL)","Isquiotibiales","Gluteos, Espalda Baja","Barra","Bisagra de cadera","Intermedio","Rodillas casi bloqueadas, maximo estiramiento de isquiotibiales."],
["Peso Muerto Sumo con Mancuernas","Gluteos","Isquiotibiales, Aductores","Mancuernas","Bisagra de cadera amplia","Principiante","Postura sumo, mancuernas frente al cuerpo."],
["Peso Muerto con Deficit","Espalda Baja","Gluteos, Isquiotibiales","Barra","Bisagra de cadera desde elevacion","Avanzado","De pie sobre una plataforma elevada, mayor rango de movimiento."],
["Peso Muerto a Una Pierna con Barra","Isquiotibiales","Gluteos, Core","Barra","Bisagra de cadera unilateral","Avanzado","Mayor demanda de equilibrio y estabilizacion."],
["Buenos Dias (Good Morning)","Espalda Baja","Isquiotibiales, Gluteos","Barra","Bisagra de cadera","Avanzado","Barra en trapecios, inclinarse hasta la horizontal manteniendo espalda neutra."],
["Hiperextension con Barra","Espalda Baja","Gluteos, Isquiotibiales","Barra","Extension de cadera/columna","Intermedio","En banco romano, sostener una barra frente al pecho."],
# Barra - remos
["Remo con Barra Prono (Pendlay)","Espalda Alta","Biceps, Romboides, Trapecios","Barra","Traccion horizontal","Intermedio","Torso paralelo al suelo, tirar la barra hacia el abdomen bajo."],
["Remo con Barra Supino","Espalda Alta","Biceps","Barra","Traccion horizontal supina","Intermedio","Agarre supino, mayor activacion del dorsal."],
["Remo con Barra Agarre Estrecho","Espalda Alta","Biceps, Romboides","Barra","Traccion horizontal estrecha","Intermedio","Manos juntas, mayor enfasis en romboides."],
["Remo al Menton con Barra","Trapecios","Hombros (lateral), Biceps","Barra","Traccion vertical","Intermedio","Agarre estrecho, tirar la barra hacia el menton."],
["Remo con Barra en T (T-Bar Row)","Espalda Alta","Biceps, Romboides, Trapecios","Barra","Traccion horizontal","Intermedio","Un extremo de la barra anclado, tirar del otro extremo con mangos."],
["Remo Chest Supported en T-Bar","Espalda Alta","Biceps, Romboides","Barra","Traccion horizontal apoyada","Intermedio","Pecho apoyado en pad inclinado, elimina el uso de la espalda baja."],
["Seal Row","Espalda Alta","Biceps, Romboides","Barra","Traccion horizontal tumbado","Avanzado","Tumbado boca abajo sobre un banco alto, remo colgando libremente."],
# Maquinas espalda
["Jalon al Pecho con Barra Amplia","Dorsal","Biceps, Romboides","Maquina","Traccion vertical amplia","Principiante","Barra larga, agarre prono, tirar hacia la clavicula."],
["Jalon al Pecho con Barra Estrecha","Dorsal","Biceps","Maquina","Traccion vertical estrecha","Principiante","Barra estrecha o mangos V, mayor rango de movimiento."],
["Jalon Tras Nuca","Dorsal","Biceps, Romboides","Maquina","Traccion vertical posterior","Intermedio","Barra detras de la cabeza, requiere buena movilidad cervical."],
["Jalon Agarre Supino","Dorsal","Biceps","Maquina","Traccion vertical supina","Principiante","Palmas mirando hacia vos, mayor activacion de biceps."],
["Remo en Maquina con Pecho Apoyado","Espalda Alta","Biceps, Romboides","Maquina","Traccion horizontal apoyada","Principiante","Pecho contra el pad, movimiento puro de espalda."],
["Remo en Maquina Sentado (Polea)","Espalda Alta","Biceps, Romboides, Trapecios","Maquina","Traccion horizontal sentado","Principiante","Pies en los pedales, tirar con mangos hacia el abdomen."],
["Remo Hammer Strength Un Brazo","Espalda Alta","Biceps","Maquina","Traccion horizontal unilateral","Principiante","Palanca articulada, un brazo a la vez."],
["Remo Hammer Strength Bilateral","Espalda Alta","Biceps, Romboides","Maquina","Traccion horizontal bilateral","Principiante","Ambos brazos simultaneamente."],
["Hiperextensiones en Banco Romano","Espalda Baja","Gluteos, Isquiotibiales","Maquina","Extension de cadera/columna","Principiante","Bajar el torso y subir hasta la horizontal."],
["Extension de Espalda en Maquina GHD","Espalda Baja","Gluteos, Isquiotibiales","Maquina","Extension de cadera/columna","Intermedio","GHD (Glute-Ham Developer), mayor rango que el banco romano."],
["Pullover en Maquina (Nautilus)","Dorsal","Triceps largo, Pecho","Maquina","Tiracion vertical/anterior","Intermedio","Codos sobre el pad, llevar los brazos de arriba hacia adelante y abajo."],
["Maquina de Remo Independiente (Plate Loaded)","Espalda Alta","Biceps, Romboides","Maquina","Traccion horizontal articulada","Intermedio","Palancas independientes, mayor rango y rotacion."],
# Cable espalda
["Jalon al Pecho (Cable Barra)","Dorsal","Biceps, Romboides","Cable","Traccion vertical","Principiante","Barra larga, agarre prono a la anchura de hombros."],
["Jalon Unilateral (Cable)","Dorsal","Biceps","Cable","Traccion vertical unilateral","Intermedio","Un brazo a la vez, mayor rango y contraccion."],
["Remo con Cable Sentado (Mangos V)","Espalda Alta","Biceps, Romboides","Cable","Traccion horizontal","Principiante","Mangos V, tirar con codos pegados al cuerpo."],
["Remo con Cable Agarre Amplio","Espalda Alta","Romboides, Trapecios","Cable","Traccion horizontal amplia","Intermedio","Barra larga, codos hacia afuera, mayor trabajo de romboides."],
["Remo Unilateral con Cable","Espalda Alta","Biceps","Cable","Traccion horizontal unilateral","Intermedio","Un brazo, mayor rango de rotacion escapular."],
["Face Pull","Trapecios","Hombros (posterior), Manguito Rotador","Cable","Traccion diagonal","Principiante","Polea alta, tirar hacia la cara separando las manos."],
["Pullover con Cable","Dorsal","Triceps largo, Pecho","Cable","Tiracion vertical/anterior","Intermedio","Polea alta, tirar hacia los muslos manteniendo brazos semi-extendidos."],
["Remo de Pie con Cable (Polea Baja)","Espalda Alta","Biceps, Romboides","Cable","Traccion horizontal de pie","Principiante","Polea baja, tirar hacia el abdomen de pie."],
["Straight Arm Pulldown","Dorsal","Triceps largo","Cable","Tiracion vertical","Intermedio","Brazos casi extendidos, tirar la barra/cuerda de arriba hacia los muslos."],
# Peso corporal espalda
["Dominadas Agarre Prono (Pull-ups)","Dorsal","Biceps, Romboides","Peso Corporal","Traccion vertical","Avanzado","Agarre prono mas ancho que los hombros, subir hasta la barbilla."],
["Dominadas Agarre Supino (Chin-ups)","Dorsal","Biceps","Peso Corporal","Traccion vertical supina","Intermedio","Agarre supino, mayor activacion de biceps."],
["Dominadas Agarre Neutro","Dorsal","Biceps, Braquial","Peso Corporal","Traccion vertical neutra","Intermedio","Manos enfrentadas, posicion intermedia."],
["Dominadas Agarre Estrecho","Dorsal","Biceps","Peso Corporal","Traccion vertical estrecha","Avanzado","Manos juntas, mayor enfasis en la parte interna del dorsal."],
["Dominadas con Lastre","Dorsal","Biceps, Romboides","Peso Corporal","Traccion vertical con carga extra","Avanzado","Cinto con disco o chaleco lastrado."],
["Dominadas Explosivas (Muscle-up)","Dorsal","Triceps, Hombros, Core","Peso Corporal","Traccion + empuje vertical","Avanzado","Subir por encima de la barra llevando el pecho."],
["Remo Invertido en Barra Baja","Espalda Alta","Biceps, Romboides","Peso Corporal","Traccion horizontal","Principiante","Cuerpo inclinado bajo la barra, tirar el pecho hacia la barra."],
["Remo Invertido TRX","Espalda Alta","Biceps, Romboides","TRX / Suspension","Traccion horizontal","Principiante","Mayor inclinacion = mayor dificultad."],
["Dead Hang","Espalda Alta","Hombros, Antebrazo","Peso Corporal","Descompresion/grip isometrico","Principiante","Colgar de la barra, descomprimir la columna, trabajar el agarre."],
["Active Hang (Escapular Pull-up)","Dorsal","Trapecios","Peso Corporal","Depresion escapular","Intermedio","Sin doblar los codos, retraer y deprimir las escapulas desde el hang."],

# ══════════════════════════════════════════════════════════════════════════════
# HOMBROS
# ══════════════════════════════════════════════════════════════════════════════
["Press Militar con Barra De Pie","Hombros","Triceps, Trapecios, Core","Barra","Empuje vertical","Intermedio","Agarre a la anchura de hombros, press sobre la cabeza."],
["Press Militar con Barra Sentado","Hombros","Triceps, Trapecios","Barra","Empuje vertical","Principiante","Version sentada, mas estable."],
["Push Press con Barra","Hombros","Triceps, Cuadriceps, Core","Barra","Empuje vertical explosivo","Intermedio","Ligero impulso de piernas para dar inercia a la barra."],
["Push Jerk","Hombros","Triceps, Cuadriceps, Core","Barra","Empuje vertical olimpico","Avanzado","Impulso de piernas, caer bajo la barra con brazos extendidos."],
["Press con Barra Landmine (Un Brazo)","Hombros","Triceps, Core","Barra","Empuje semi-vertical","Intermedio","Un extremo de la barra anclado en esquina, press diagonal."],
["Press con Barra Landmine (Bilateral)","Hombros","Triceps, Pecho","Barra","Empuje semi-vertical bilateral","Intermedio","Ambas manos en la barra, enfasis pecho-hombro."],
["Remo al Menton con Mancuernas","Trapecios","Hombros (lateral), Biceps","Mancuernas","Traccion vertical","Intermedio","Mancuernas frente al cuerpo, codos por encima de las munecas."],
["Press de Hombros con Mancuernas Sentado","Hombros","Triceps","Mancuernas","Empuje vertical","Principiante","Banco con respaldo, mas estable."],
["Press de Hombros con Mancuernas De Pie","Hombros","Triceps, Core","Mancuernas","Empuje vertical de pie","Intermedio","Mayor demanda de estabilizacion del core."],
["Elevaciones Laterales","Hombros (lateral)","Trapecios","Mancuernas","Aislamiento lateral","Principiante","Brazos ligeramente flexionados, elevar hasta la altura del hombro."],
["Elevaciones Laterales Inclinado","Hombros (lateral)","Trapecios","Mancuernas","Aislamiento lateral inclinado","Intermedio","Cuerpo inclinado lateralmente, mayor rango efectivo."],
["Elevaciones Laterales Acostado","Hombros (lateral)","Trapecios","Mancuernas","Aislamiento lateral tumbado","Intermedio","Tumbado de lado en banco, elimina el balanceo."],
["Elevaciones Frontales Alternadas","Hombros (anterior)","Pecho Superior","Mancuernas","Aislamiento frontal","Principide","Alternar brazos, elevar al frente hasta la horizontal."],
["Elevaciones Frontales Simultaneas","Hombros (anterior)","Pecho Superior","Mancuernas","Aislamiento frontal bilateral","Principiante","Ambos brazos a la vez."],
["Pajaro / Elevaciones Posteriores","Hombros (posterior)","Romboides, Trapecios","Mancuernas","Aislamiento posterior","Principiante","Inclinado hacia adelante, abrir los brazos como alas."],
["Press Arnold","Hombros","Triceps, Pecho","Mancuernas","Empuje vertical con rotacion","Intermedio","Palmas hacia vos al inicio, rotar y extender sobre la cabeza."],
["Press Cubano con Mancuernas","Hombros (posterior)","Manguito Rotador, Trapecios","Mancuernas","Rotacion externa + empuje","Intermedio","Codos a 90, rotar los antebrazos hacia arriba, luego presionar."],
["Rotacion Externa con Mancuerna","Manguito Rotador","Hombros (posterior)","Mancuernas","Rotacion externa de hombro","Principiante","Codo a 90 apoyado en el muslo o banco, rotar externamente el antebrazo."],
["Rotacion Interna con Mancuerna","Manguito Rotador","Hombros (anterior)","Mancuernas","Rotacion interna de hombro","Principiante","Codo a 90, rotar internamente el antebrazo."],
["W's con Mancuernas (Prone W)","Hombros (posterior)","Manguito Rotador, Romboides","Mancuernas","Rotacion externa tumbado","Principiante","Tumbado boca abajo en banco, brazos en forma de W, rotar hacia arriba."],
["Press de Hombros en Maquina","Hombros","Triceps","Maquina","Empuje vertical guiado","Principiante","Movimiento guiado, ideal para aprender."],
["Elevaciones Laterales en Maquina","Hombros (lateral)","Trapecios","Maquina","Aislamiento lateral guiado","Principiante","Tension constante a lo largo del rango."],
["Reverse Pec Deck","Hombros (posterior)","Romboides","Maquina","Aislamiento posterior","Principiante","Codos casi extendidos, abrir hacia atras."],
["Maquina de Hombros Hammer Strength","Hombros","Triceps","Maquina","Empuje vertical articulado","Principiante","Palanca articulada, movimiento mas natural."],
["Elevaciones Laterales con Cable","Hombros (lateral)","Trapecios","Cable","Aislamiento lateral","Principiante","Tension constante a diferencia de la mancuerna."],
["Elevaciones Laterales Unilateral Cable","Hombros (lateral)","Trapecios","Cable","Aislamiento lateral unilateral","Principiante","Polea baja, un brazo, tension constante."],
["Elevaciones Frontales con Cable","Hombros (anterior)","Pecho Superior","Cable","Aislamiento frontal","Principiante","Polea baja, elevar hacia el frente."],
["Elevaciones Posteriores con Cable","Hombros (posterior)","Romboides","Cable","Aislamiento posterior","Principiante","Polea alta, cruzando el cuerpo hacia abajo."],
["Press Vertical con Cable (Un Brazo)","Hombros","Triceps","Cable","Empuje vertical unilateral","Intermedio","Polea baja, press sobre la cabeza unilateral."],
["Rotacion Externa con Cable","Manguito Rotador","Hombros (posterior)","Cable","Rotacion externa","Principiante","Polea baja, codo a 90 pegado al cuerpo, rotar externamente."],
["Rotacion Interna con Cable","Manguito Rotador","Hombros (anterior)","Cable","Rotacion interna","Principiante","Polea baja, codo a 90, rotar hacia el abdomen."],
["Handstand Push-up (Contra la Pared)","Hombros","Triceps, Trapecios","Peso Corporal","Empuje vertical invertido","Avanzado","Parado de manos contra la pared, bajar la cabeza al suelo y empujar."],
["Pike Push-up","Hombros","Triceps","Peso Corporal","Empuje vertical","Intermedio","Caderas elevadas, cuerpo en V invertida, bajar la cabeza al suelo."],

# ══════════════════════════════════════════════════════════════════════════════
# BICEPS
# ══════════════════════════════════════════════════════════════════════════════
["Curl de Biceps con Barra","Biceps","Braquial, Braquiorradial","Barra","Flexion de codo","Principiante","Codos fijos a los costados, curl completo."],
["Curl de Biceps con Barra EZ","Biceps","Braquial, Braquiorradial","Barra","Flexion de codo","Principiante","Agarre semisupino, menor estres en las munecas."],
["Curl con Barra EZ Agarre Invertido","Braquiorradial","Biceps, Extensores del Antebrazo","Barra","Flexion de codo prono","Principiante","Agarre prono, enfasis en braquiorradial."],
["Curl 21s con Barra","Biceps","Braquial","Barra","Flexion de codo parcial","Intermedio","7 reps en la mitad inferior, 7 en la superior, 7 completas."],
["Curl de Biceps en Banco Scott con Barra EZ","Biceps","Braquial","Barra","Flexion de codo apoyada","Principiante","Brazo apoyado en el pad inclinado, elimina el balanceo."],
["Curl de Biceps con Barra Landmine","Biceps","Braquial","Barra","Flexion de codo diagonal","Intermedio","Un extremo de la barra anclado, curl en plano diagonal."],
["Curl Alternado con Mancuernas","Biceps","Braquial","Mancuernas","Flexion de codo unilateral","Principiante","Un brazo a la vez, supinar al subir."],
["Curl Simultaneo con Mancuernas","Biceps","Braquial","Mancuernas","Flexion de codo bilateral","Principiante","Ambos brazos a la vez."],
["Curl Martillo","Braquiorradial","Biceps, Braquial","Mancuernas","Flexion de codo neutral","Principiante","Agarre neutro, pulgares arriba."],
["Curl Martillo Cruzado","Braquiorradial","Biceps, Braquial","Mancuernas","Flexion de codo cruzada","Principiante","Elevar la mancuerna cruzando frente al pecho."],
["Curl Concentrado","Biceps","Braquial","Mancuernas","Flexion de codo aislada","Principiante","Codo apoyado en cara interna del muslo, maximo aislamiento."],
["Curl en Banco Scott con Mancuernas","Biceps","Braquial","Mancuernas","Flexion de codo apoyada","Principiante","Brazo apoyado en pad inclinado."],
["Curl Spider","Biceps","Braquial","Mancuernas","Flexion de codo inclinada","Intermedio","Pecho apoyado en banco inclinado, codos apuntando al suelo."],
["Curl Inclinado Stretch","Biceps","Braquial","Mancuernas","Flexion de codo elongada","Intermedio","Banco inclinado, maximo estiramiento del biceps."],
["Curl de Muneca Supino con Mancuerna","Flexores del Antebrazo","Biceps","Mancuernas","Flexion de muneca","Principiante","Antebrazo apoyado, solo mueve la muneca."],
["Curl Zottman","Biceps","Braquiorradial, Extensores del Antebrazo","Mancuernas","Flexion + rotacion de codo","Intermedio","Subir en supinacion, bajar en pronacion, trabaja ambas superficies."],
["Curl de Biceps en Maquina Sentado","Biceps","Braquial","Maquina","Flexion de codo guiada","Principiante","Movimiento guiado, tension constante."],
["Curl de Biceps en Maquina de Predicador","Biceps","Braquial","Maquina","Flexion de codo apoyada guiada","Principiante","Brazo sobre pad inclinado, movimiento guiado."],
["Curl de Biceps en Cable Polea Baja","Biceps","Braquial","Cable","Flexion de codo","Principiante","Tension constante en todo el rango."],
["Curl Unilateral en Cable","Biceps","Braquial","Cable","Flexion de codo unilateral","Principiante","Un brazo, maximo rango."],
["Curl de Biceps Polea Alta (Supino)","Biceps","Braquial","Cable","Flexion de codo en abduccion","Intermedio","Brazos abiertos a altura del hombro, doblar los codos."],
["Curl en Cable con Barra EZ","Biceps","Braquial","Cable","Flexion de codo","Principiante","Barra EZ en polea baja, igual que barra pero con tension constante."],
["Curl en TRX (Suspension)","Biceps","Core","TRX / Suspension","Flexion de codo suspendida","Intermedio","Cuerpo inclinado, curl hacia la frente usando el peso corporal."],
["Chin-ups (Agarre Supino)","Biceps","Dorsal","Peso Corporal","Traccion vertical","Intermedio","Agarre supino, mayor activacion de biceps."],
["Curl de Biceps con Banda","Biceps","Braquial","Bandas Elasticas","Flexion de codo","Principiante","Banda bajo los pies, curl completo."],

# ══════════════════════════════════════════════════════════════════════════════
# TRICEPS
# ══════════════════════════════════════════════════════════════════════════════
["Press Frances con Barra EZ (Skullcrusher)","Triceps","","Barra","Extension de codo acostado","Intermedio","Barra a la frente doblando solo los codos."],
["Press Frances Inclinado con Barra EZ","Triceps","","Barra","Extension de codo inclinada","Intermedio","Banco inclinado, mayor rango de movimiento."],
["Press de Banca Agarre Estrecho","Triceps","Pecho, Hombros","Barra","Empuje estrecho","Intermedio","Codos pegados al cuerpo, enfasis en triceps."],
["JM Press","Triceps","Pecho","Barra","Hibrido press-skullcrusher","Avanzado","Hibrido entre press estrecho y press frances, muy eficiente para triceps."],
["Press Frances con Mancuernas","Triceps","","Mancuernas","Extension de codo acostado","Principiante","Mancuernas bajan a los costados de la cabeza."],
["Extension de Triceps Sobre la Cabeza","Triceps","","Mancuernas","Extension de codo sobre cabeza","Principiante","Mancuerna detras de la cabeza, extender los codos."],
["Extension de Triceps Sobre Cabeza Bilateral","Triceps","","Mancuernas","Extension de codo sobre cabeza bilateral","Principiante","Una mancuerna sostenida con ambas manos detras de la cabeza."],
["Patada de Triceps (Kickback)","Triceps","","Mancuernas","Extension de codo","Principiante","Inclinado, codo fijo, extender el brazo hacia atras."],
["Pushdown de Triceps con Barra","Triceps","","Cable","Extension de codo","Principiante","Polea alta, codos fijos a los costados."],
["Pushdown de Triceps con Cuerda","Triceps","","Cable","Extension de codo con rotacion","Principiante","Al final separar la cuerda hacia afuera."],
["Pushdown de Triceps con Barra V","Triceps","","Cable","Extension de codo","Principiante","Barra V, agarre comodo para las munecas."],
["Pushdown Unilateral con Cable","Triceps","","Cable","Extension de codo unilateral","Principiante","Un brazo a la vez, mayor rango."],
["Extension de Triceps Sobre Cabeza (Cable)","Triceps","","Cable","Extension de codo sobre cabeza","Principiante","Polea alta, manos detras de la cabeza."],
["Extension de Triceps Sobre Cabeza con Cuerda","Triceps","","Cable","Extension de codo sobre cabeza","Principiante","Polea alta, cuerda, mayor rango de separacion."],
["Press Frances en Cable","Triceps","","Cable","Extension de codo acostado","Intermedio","Tumbado frente a la polea baja, skullcrusher con cable."],
["Extension Triceps en Maquina","Triceps","","Maquina","Extension de codo guiada","Principiante","Tension constante, movimiento guiado."],
["Dips en Maquina Asistida","Triceps","Pecho, Hombros","Maquina","Empuje vertical asistido","Principiante","Maquina contrarresta parte del peso corporal."],
["Fondos en Paralelas (Triceps)","Triceps","Pecho, Hombros","Peso Corporal","Empuje vertical","Intermedio","Cuerpo erguido, codos pegados."],
["Fondos en Banco (Bench Dips)","Triceps","Pecho","Peso Corporal","Empuje decline","Principiante","Manos en banco detras, pies en suelo o elevados."],
["Fondos en Banco Pies Elevados","Triceps","Pecho","Peso Corporal","Empuje decline avanzado","Intermedio","Pies sobre otro banco, mayor carga."],
["Diamond Push-up","Triceps","Pecho","Peso Corporal","Empuje estrecho","Intermedio","Manos en diamante bajo el pecho."],
["Extension de Triceps en Anillas","Triceps","Core, Hombros","Peso Corporal","Extension de codo en anillas","Avanzado","Anillas, mayor inestabilidad y rango."],
["Extension de Triceps Sobre Cabeza con Banda","Triceps","","Bandas Elasticas","Extension de codo sobre cabeza","Principiante","Banda anclada abajo o pisada, extension sobre la cabeza."],

# ══════════════════════════════════════════════════════════════════════════════
# CUADRICEPS
# ══════════════════════════════════════════════════════════════════════════════
["Sentadilla con Barra Alta (High Bar)","Cuadriceps","Gluteos, Isquiotibiales, Core","Barra","Sentadilla","Intermedio","Barra en trapecios superior, torso mas erguido, mayor profundidad."],
["Sentadilla con Barra Baja (Low Bar)","Cuadriceps","Gluteos, Isquiotibiales, Espalda Baja","Barra","Sentadilla","Avanzado","Barra en trapecios medios, mayor inclinacion del torso."],
["Sentadilla Frontal","Cuadriceps","Gluteos, Core","Barra","Sentadilla","Avanzado","Barra en hombros delanteros, torso muy erguido."],
["Sentadilla de Zercher","Cuadriceps","Gluteos, Core, Biceps","Barra","Sentadilla","Avanzado","Barra en el pliegue de los codos, postura muy erguida."],
["Sentadilla Overhead (OHS)","Cuadriceps","Core, Hombros, Gluteos","Barra","Sentadilla con barra sobre cabeza","Avanzado","Requiere gran movilidad y estabilidad general."],
["Sentadilla Sumo con Barra","Gluteos","Cuadriceps, Aductores","Barra","Sentadilla amplia","Intermedio","Pies muy abiertos, puntas hacia afuera."],
["Sentadilla con Bandas (Barra)","Cuadriceps","Gluteos","Barra","Sentadilla con resistencia variable","Intermedio","Bandas ancladas al suelo o al rack."],
["Sentadilla con Cadenas (Barra)","Cuadriceps","Gluteos","Barra","Sentadilla con resistencia variable","Avanzado","Cadenas colgadas de la barra aumentan la carga al subir."],
["Good Morning (sentadilla estilo)","Cuadriceps","Isquiotibiales, Gluteos, Espalda Baja","Barra","Bisagra/sentadilla","Avanzado","Mas inclinado que el gm clasico, enfasis en cuadriceps."],
["Zancadas con Barra Al Frente","Cuadriceps","Gluteos, Isquiotibiales","Barra","Zancada hacia adelante","Intermedio","Un paso al frente, doblar ambas rodillas a 90."],
["Zancadas con Barra Al Atras","Cuadriceps","Gluteos","Barra","Zancada hacia atras","Intermedio","Paso hacia atras, mayor control de rodilla."],
["Zancadas Caminando con Barra","Cuadriceps","Gluteos, Isquiotibiales","Barra","Zancada caminando","Intermedio","Avanzar con pasos alternos."],
["Sentadilla Goblet con Mancuerna","Cuadriceps","Gluteos, Core","Mancuernas","Sentadilla","Principiante","Mancuerna frente al pecho, torso erguido."],
["Zancadas con Mancuernas Al Frente","Cuadriceps","Gluteos, Isquiotibiales","Mancuernas","Zancada hacia adelante","Principiante","Paso al frente, rodillas a 90."],
["Zancadas Inversas con Mancuernas","Cuadriceps","Gluteos","Mancuernas","Zancada hacia atras","Principiante","Paso hacia atras, mayor control de rodilla."],
["Zancadas Laterales con Mancuernas","Gluteos","Cuadriceps, Aductores","Mancuernas","Zancada lateral","Intermedio","Paso lateral, doblar la rodilla que carga el peso."],
["Zancadas Caminando con Mancuernas","Cuadriceps","Gluteos, Isquiotibiales","Mancuernas","Zancada caminando","Principiante","Avanzar con pasos alternos."],
["Step-Up con Mancuernas","Cuadriceps","Gluteos","Mancuernas","Unilateral cadera/rodilla","Principiante","Subir a un banco con una pierna a la vez."],
["Extension de Cuadriceps en Maquina","Cuadriceps","","Maquina","Extension de rodilla","Principiante","Sentado, extender las rodillas."],
["Extension Unilateral en Maquina","Cuadriceps","","Maquina","Extension de rodilla unilateral","Principiante","Una pierna a la vez, mayor concentracion."],
["Prensa de Piernas (Leg Press) 45","Cuadriceps","Gluteos, Isquiotibiales","Maquina","Empuje bilateral 45","Principiante","Angulo 45, pies en la plataforma."],
["Prensa de Piernas Horizontal","Cuadriceps","Gluteos, Isquiotibiales","Maquina","Empuje bilateral horizontal","Principiante","Angulo horizontal, diferente tension en la articulacion."],
["Prensa de Piernas Unilateral","Cuadriceps","Gluteos","Maquina","Empuje unilateral","Intermedio","Una pierna a la vez, detecta desequilibrios."],
["Hack Squat en Maquina","Cuadriceps","Gluteos","Maquina","Sentadilla guiada","Principiante","Espalda contra el pad inclinado."],
["Pendulum Squat","Cuadriceps","Gluteos","Maquina","Sentadilla en arco","Intermedio","Maquina de arco pendular, excelente para cuadriceps sin carga en la columna."],
["Belt Squat (Sentadilla con Cinto)","Cuadriceps","Gluteos","Maquina","Sentadilla sin carga en columna","Intermedio","La carga se cuelga del cinto en la cadera, sin carga en columna."],
["Sissy Squat","Cuadriceps","","Maquina","Extension de rodilla de pie","Avanzado","Inclinarse hacia atras mientras las rodillas avanzan hacia adelante."],
["Sentadilla en Maquina Smith","Cuadriceps","Gluteos","Maquina Smith","Sentadilla guiada","Principiante","Barra guiada."],
["Sentadilla Libre (Sin peso)","Cuadriceps","Gluteos, Core","Peso Corporal","Sentadilla","Principiante","Pies a la anchura de los hombros."],
["Sentadilla Bulgara","Cuadriceps","Gluteos, Isquiotibiales","Peso Corporal","Zancada dividida","Intermedio","Pie trasero elevado en banco."],
["Pistol Squat","Cuadriceps","Gluteos, Core","Peso Corporal","Sentadilla unilateral","Avanzado","Sentadilla con una sola pierna."],
["Wall Sit (Sentadilla en Pared)","Cuadriceps","","Peso Corporal","Isometria de cuadriceps","Principiante","Espalda contra la pared, muslos paralelos al suelo."],
["Box Squat","Cuadriceps","Gluteos, Isquiotibiales","Peso Corporal","Sentadilla a caja","Intermedio","Sentarse sobre una caja/banco y volver a subir."],
["Jump Squat","Cuadriceps","Gluteos, Core","Peso Corporal","Sentadilla explosiva","Intermedio","Sentadilla profunda y saltar explosivamente."],
["Sentadilla con Banda","Cuadriceps","Gluteos","Bandas Elasticas","Sentadilla","Principiante","Banda sobre los hombros o bajo los pies."],

# ══════════════════════════════════════════════════════════════════════════════
# ISQUIOTIBIALES
# ══════════════════════════════════════════════════════════════════════════════
["Peso Muerto Sumo (Isquios)","Isquiotibiales","Gluteos, Aductores, Cuadriceps","Barra","Bisagra de cadera amplia","Avanzado","Postura sumo, pies muy abiertos."],
["Peso Muerto Rumano con Barra","Isquiotibiales","Gluteos, Espalda Baja","Barra","Bisagra de cadera","Intermedio","Rodillas ligeramente flexionadas, barra baja por las piernas."],
["Peso Muerto Piernas Rigidas","Isquiotibiales","Gluteos, Espalda Baja","Barra","Bisagra de cadera","Intermedio","Rodillas casi bloqueadas, maximo estiramiento."],
["Buenos Dias con Barra","Isquiotibiales","Gluteos, Espalda Baja","Barra","Bisagra de cadera","Avanzado","Barra en trapecios, inclinarse hasta la horizontal."],
["Peso Muerto Rumano con Mancuernas","Isquiotibiales","Gluteos, Espalda Baja","Mancuernas","Bisagra de cadera","Principiante","Igual que barra pero con mancuernas."],
["Peso Muerto Una Pierna con Mancuerna","Isquiotibiales","Gluteos, Core","Mancuernas","Bisagra de cadera unilateral","Intermedio","Equilibrio y estabilizacion unilateral."],
["Curl de Piernas Tumbado","Isquiotibiales","Gastrocnemio","Maquina","Flexion de rodilla","Principiante","Tumbado boca abajo, curvar los talones hacia los gluteos."],
["Curl de Piernas Sentado","Isquiotibiales","","Maquina","Flexion de rodilla","Principiante","Sentado, tension constante en todo el rango."],
["Curl de Piernas De Pie Unilateral","Isquiotibiales","","Maquina","Flexion de rodilla unilateral","Principiante","De pie, una pierna a la vez."],
["Curl de Piernas en Maquina GHD","Isquiotibiales","Gluteos","Maquina","Flexion de rodilla + extension cadera","Avanzado","En GHD, bajar y subir activando gluteo e isquio."],
["Nordic Curl","Isquiotibiales","","Peso Corporal","Flexion de rodilla excentrica","Avanzado","Talones sujetos, descender lentamente controlando la caida."],
["Curl de Isquios con Pelota de Estabilidad","Isquiotibiales","Gluteos, Core","Peso Corporal","Flexion de rodilla desde el suelo","Intermedio","Tumbado, pies sobre pelota, elevar caderas y curvar las rodillas."],
["Good Morning con Bandas","Isquiotibiales","Gluteos, Espalda Baja","Bandas Elasticas","Bisagra de cadera","Intermedio","Banda sobre los hombros anclada al suelo, bisagra de cadera."],
["Peso Muerto Rumano con Cable","Isquiotibiales","Gluteos","Cable","Bisagra de cadera","Principiante","Tobillera o agarre bajo, bisagra con cable."],
["Curl de Isquios con Cable (Tumbado)","Isquiotibiales","","Cable","Flexion de rodilla","Principiante","Tobillera, tumbado, curvar la rodilla hacia el gluteo."],
["Curl de Isquios con Cable (De Pie)","Isquiotibiales","","Cable","Flexion de rodilla de pie","Principiante","Tobillera, de pie, curvar la rodilla."],

# ══════════════════════════════════════════════════════════════════════════════
# GLUTEOS
# ══════════════════════════════════════════════════════════════════════════════
["Hip Thrust con Barra","Gluteos","Isquiotibiales, Cuadriceps","Barra","Extension de cadera","Intermedio","Espalda en banco, barra en cadera, empujar las caderas."],
["Hip Thrust con Barra Pausa","Gluteos","Isquiotibiales","Barra","Extension de cadera con pausa","Avanzado","Pausar 2 seg arriba para maximizar la contraccion."],
["Sentadilla Sumo con Barra","Gluteos","Cuadriceps, Aductores, Isquiotibiales","Barra","Sentadilla amplia","Intermedio","Pies muy abiertos, puntas hacia afuera."],
["Hip Thrust con Mancuerna","Gluteos","Isquiotibiales","Mancuernas","Extension de cadera","Principiante","Mancuerna sobre la cadera."],
["Hip Thrust Unilateral con Mancuerna","Gluteos","Isquiotibiales, Core","Mancuernas","Extension de cadera unilateral","Intermedio","Una pierna a la vez."],
["Peso Muerto Sumo con Mancuernas","Gluteos","Isquiotibiales, Aductores","Mancuernas","Bisagra de cadera amplia","Principiante","Postura sumo, mancuernas frente al cuerpo."],
["Abduccion de Cadera en Maquina","Gluteos (medio/menor)","TFL","Maquina","Abduccion de cadera sentado","Principiante","Separar las piernas contra la resistencia."],
["Aduccion de Cadera en Maquina","Aductores","Gluteos","Maquina","Aduccion de cadera sentado","Principiante","Juntar las piernas contra la resistencia."],
["Extension de Cadera en Maquina","Gluteos","Isquiotibiales","Maquina","Extension de cadera","Principiante","De pie o inclinado, extender la pierna hacia atras."],
["Hip Thrust en Maquina","Gluteos","Isquiotibiales","Maquina","Extension de cadera guiada","Principiante","Maquina especifica de hip thrust."],
["Reverse Hyper (Extension Inversa)","Gluteos","Isquiotibiales, Espalda Baja","Maquina","Extension de cadera invertida","Intermedio","Tumbado boca abajo, elevar las piernas hacia atras."],
["Cable Kickback","Gluteos","Isquiotibiales","Cable","Extension de cadera","Principiante","Tobillera, polea baja, extender la pierna hacia atras."],
["Abduccion de Cadera con Cable","Gluteos (medio)","","Cable","Abduccion de cadera","Principiante","Tobillera, elevar la pierna lateralmente."],
["Hip Thrust con Cable","Gluteos","Isquiotibiales","Cable","Extension de cadera con cable","Intermedio","Cable entre las piernas, empujar las caderas hacia arriba."],
["Puente de Gluteos","Gluteos","Isquiotibiales, Core","Peso Corporal","Extension de cadera","Principiante","Tumbado, pies apoyados, elevar las caderas."],
["Puente de Gluteos Unilateral","Gluteos","Isquiotibiales, Core","Peso Corporal","Extension de cadera unilateral","Intermedio","Una pierna extendida."],
["Hip Thrust con Peso Corporal","Gluteos","Isquiotibiales","Peso Corporal","Extension de cadera","Principiante","Espalda en banco, sin peso adicional."],
["Donkey Kicks","Gluteos","","Peso Corporal","Extension de cadera en cuadrupedia","Principiante","En cuatro apoyos, llevar el talon hacia el techo."],
["Fire Hydrant","Gluteos (medio)","","Peso Corporal","Abduccion de cadera en cuadrupedia","Principiante","En cuatro apoyos, elevar la rodilla hacia afuera como un perro."],
["Clamshell con Banda","Gluteos (medio)","","Bandas Elasticas","Rotacion externa de cadera","Principiante","Tumbado de lado, rodillas juntas, abrir como una almeja."],
["Monster Walk","Gluteos (medio)","TFL","Bandas Elasticas","Abduccion dinamica","Principiante","Banda en rodillas, pasos laterales."],
["Hip Abduction con Banda de Pie","Gluteos (medio)","","Bandas Elasticas","Abduccion de cadera de pie","Principiante","Banda en tobillos, elevar la pierna lateralmente."],
["Hip Thrust con Banda","Gluteos","Isquiotibiales","Bandas Elasticas","Extension de cadera con banda","Principiante","Banda sobre la cadera."],
["Sentadilla Sumo con Banda","Gluteos","Cuadriceps, Aductores","Bandas Elasticas","Sentadilla amplia con banda","Principiante","Banda sobre los hombros o bajo los pies."],
["Frog Pumps","Gluteos","","Peso Corporal","Extension de cadera rango corto","Principiante","Tumbado, plantas de los pies juntas, elevar las caderas en rango corto."],
["Step-Up con Mancuernas","Gluteos","Cuadriceps","Mancuernas","Unilateral","Principiante","Subir a un banco con una pierna a la vez."],

# ══════════════════════════════════════════════════════════════════════════════
# PANTORRILLAS
# ══════════════════════════════════════════════════════════════════════════════
["Elevacion de Talones De Pie con Barra","Gastrocnemio","Soleo","Barra","Flexion plantar","Principiante","Barra en trapecios, elevarse sobre los dedos."],
["Elevacion de Talones con Mancuernas","Gastrocnemio","Soleo","Mancuernas","Flexion plantar","Principiante","Mancuernas a los costados."],
["Elevacion de Talones De Pie en Maquina","Gastrocnemio","Soleo","Maquina","Flexion plantar","Principiante","Hombros bajo los pads."],
["Elevacion de Talones Sentado en Maquina","Soleo","Gastrocnemio","Maquina","Flexion plantar sentado","Principiante","Rodillas a 90, enfasis en soleo."],
["Elevacion de Talones en Leg Press","Gastrocnemio","Soleo","Maquina","Flexion plantar en prensa","Principiante","Solo la parte delantera del pie en la plataforma."],
["Elevacion de Talones Unilateral De Pie","Gastrocnemio","Soleo","Peso Corporal","Flexion plantar unilateral","Intermedio","Una pierna, mayor rango y carga."],
["Elevacion de Talones Sentado (Sin Maquina)","Soleo","Gastrocnemio","Peso Corporal","Flexion plantar sentado","Principiante","Con un plato sobre los muslos si no se tiene la maquina."],
["Saltos de Pantorrilla","Gastrocnemio","","Peso Corporal","Flexion plantar explosiva","Principiante","Saltar usando solo la pantorrilla, sin doblar rodillas."],
["Elevacion de Talones con Cable","Gastrocnemio","Soleo","Cable","Flexion plantar","Principiante","Polea baja, tobillera, elevar sobre los dedos."],
["Elevacion de Talones en Smith","Gastrocnemio","Soleo","Maquina Smith","Flexion plantar","Principiante","Barra guiada, pies en una plataforma."],

# ══════════════════════════════════════════════════════════════════════════════
# TRAPECIOS
# ══════════════════════════════════════════════════════════════════════════════
["Encogimientos con Barra (Shrugs)","Trapecios","Elevador de la Escapula","Barra","Elevacion de escapula","Principiante","Barra frente al cuerpo, elevar los hombros sin rotar."],
["Encogimientos Detras de la Espalda","Trapecios","Elevador de la Escapula","Barra","Elevacion de escapula posterior","Intermedio","Barra detras del cuerpo, mayor rango posterior."],
["Remo al Menton con Barra (Trapecios)","Trapecios","Hombros (lateral), Biceps","Barra","Traccion vertical","Intermedio","Barra cerca del cuerpo, codos por encima de las munecas."],
["Encogimientos con Mancuernas","Trapecios","Elevador de la Escapula","Mancuernas","Elevacion de escapula","Principiante","Mancuernas a los costados."],
["Encogimientos con Mancuernas Unilateral","Trapecios","Elevador de la Escapula","Mancuernas","Elevacion de escapula unilateral","Principiante","Un hombro a la vez."],
["Encogimientos en Maquina Smith","Trapecios","","Maquina Smith","Elevacion de escapula guiada","Principiante","Barra guiada."],
["Encogimientos con Cable (Polea Baja)","Trapecios","","Cable","Elevacion de escapula","Principiante","Polea baja, tension constante."],
["Face Pull (Trapecios)","Trapecios","Hombros (posterior), Manguito Rotador","Cable","Traccion diagonal","Principiante","Polea alta, tirar hacia la cara separando las manos."],
["Remo al Menton con Cable","Trapecios","Hombros (lateral), Biceps","Cable","Traccion vertical","Intermedio","Polea baja, misma mecanica que con barra."],
["Encogimientos con Kettlebell","Trapecios","Elevador de la Escapula","Kettlebell","Elevacion de escapula","Principiante","Kettlebell a los costados, mismo movimiento que mancuernas."],

# ══════════════════════════════════════════════════════════════════════════════
# CORE / ABDOMEN
# ══════════════════════════════════════════════════════════════════════════════
["Plancha Frontal","Core","Hombros, Gluteos","Peso Corporal","Isometria frontal","Principiante","Antebrazos en el suelo, cuerpo recto."],
["Plancha Lateral","Oblicuos","Core, Abductores","Peso Corporal","Isometria lateral","Principiante","Un antebrazo en el suelo, cuerpo lateral recto."],
["Plancha Lateral con Elevacion de Cadera","Oblicuos","Core","Peso Corporal","Isometria lateral dinamica","Intermedio","En plancha lateral, bajar y subir la cadera."],
["Plancha con Toque de Hombro","Core","Hombros, Pecho","Peso Corporal","Antirrotacion dinamica","Intermedio","En plancha alta, tocar el hombro opuesto alternando."],
["Plancha con Extension de Brazo","Core","Hombros","Peso Corporal","Antirrotacion dinamica","Intermedio","En plancha, extender un brazo al frente y alternarlo."],
["Plancha RKC (Cuerpo Tenso)","Core","Gluteos, Cuadriceps","Peso Corporal","Isometria de alta tension","Intermedio","Plancha con maximo tension en todo el cuerpo."],
["Crunch Abdominal","Recto Abdominal","Oblicuos","Peso Corporal","Flexion de columna","Principiante","Elevar solo los hombros contrayendo el abdomen."],
["Crunch Inverso","Recto Abdominal (inferior)","Core","Peso Corporal","Flexion de cadera/columna","Principiante","Elevar la pelvis llevando rodillas al pecho."],
["Crunch Bicicleta","Oblicuos","Recto Abdominal","Peso Corporal","Rotacion + flexion","Principiante","Llevar codo hacia rodilla opuesta alternando."],
["Crunch en Polea Alta","Recto Abdominal","Oblicuos","Cable","Flexion de columna","Principiante","De rodillas frente a la polea, contraer el abdomen."],
["Crunch con Peso (Plato)","Recto Abdominal","Oblicuos","Maquina","Flexion de columna con carga","Intermedio","Sujetar un plato detras de la cabeza."],
["Elevacion de Piernas","Recto Abdominal (inferior)","Iliopsoas, Core","Peso Corporal","Flexion de cadera","Intermedio","Tumbado, elevar piernas juntas hasta 90."],
["Elevacion de Piernas Colgando (Hanging Leg Raise)","Recto Abdominal (inferior)","Iliopsoas, Core","Peso Corporal","Flexion de cadera colgando","Avanzado","Colgando de la barra, elevar piernas hasta 90 o mas."],
["Elevacion de Rodillas Colgando","Recto Abdominal (inferior)","Iliopsoas","Peso Corporal","Flexion de cadera colgando","Intermedio","Version mas facil que el de piernas estiradas."],
["Elevacion de Piernas en Silla Romana","Recto Abdominal (inferior)","Iliopsoas","Maquina","Flexion de cadera con soporte","Principiante","Antebrazos apoyados, elevar rodillas o piernas estiradas."],
["L-Sit","Core","Cuadriceps, Hombros","Peso Corporal","Isometria de flexion de cadera","Avanzado","Sentado en suelo o paralelas, elevar las piernas estiradas."],
["Dragon Flag","Core","Espalda Baja, Hombros","Peso Corporal","Isometria/Flexion de cuerpo entero","Avanzado","Tumbado en banco, bajar el cuerpo recto desde los hombros."],
["Tijeras Abdominales","Recto Abdominal","Oblicuos, Core","Peso Corporal","Flexion de cadera alternada","Principiante","Piernas extendidas, alternar arriba y abajo."],
["Mountain Climbers","Core","Hombros, Cuadriceps","Peso Corporal","Dinamico de core","Principiante","Plancha alta, llevar rodillas al pecho alternando."],
["Dead Bug","Core","Espalda Baja","Peso Corporal","Estabilizacion antiextension","Principiante","Tumbado, extender brazo y pierna opuestos."],
["Hollow Body Hold","Core","Espalda Baja","Peso Corporal","Isometria de flexion","Intermedio","Elevar hombros y piernas creando una cascara."],
["Ab Wheel Rollout","Core","Dorsal, Hombros","Peso Corporal","Antiextension dinamica","Avanzado","Rueda abdominal, extender brazos manteniendo columna neutra."],
["Ab Wheel desde Rodillas","Core","Dorsal, Hombros","Peso Corporal","Antiextension dinamica","Intermedio","Version mas facil de rodillas."],
["Pallof Press","Core","Oblicuos","Cable","Antirrotacion","Intermedio","Polea media, empujar frente al pecho resistiendo la rotacion."],
["Pallof Press con Banda","Core","Oblicuos","Bandas Elasticas","Antirrotacion","Intermedio","Banda anclada, misma mecanica que con cable."],
["Rotacion de Tronco con Cable","Oblicuos","Core","Cable","Rotacion de tronco","Principiante","Polea media, rotar el tronco."],
["Russian Twist","Oblicuos","Core","Peso Corporal","Rotacion de tronco","Principiante","Sentado inclinado, rotar el tronco de lado a lado."],
["Russian Twist con Disco","Oblicuos","Core","Maquina","Rotacion de tronco con carga","Intermedio","Sujetar un plato para mayor resistencia."],
["Windmill con Kettlebell","Oblicuos","Core, Hombros","Kettlebell","Flexion lateral + rotacion","Avanzado","Kettlebell sobre la cabeza, inclinar el torso lateralmente."],
["Flexion Lateral con Mancuerna","Oblicuos","","Mancuernas","Flexion lateral","Principiante","De pie, inclinar lateralmente."],
["Farmer Walk","Core","Trapecios, Antebrazo, Gluteos","Mancuernas","Estabilizacion dinamica de core","Principiante","Mancuernas a los costados, caminar con torso erguido."],
["Farmer Walk con Carga Asimetrica","Core","Trapecios, Oblicuos","Mancuernas","Antirrotacion/Antiflexion lateral","Intermedio","Diferente peso en cada mano, mayor demanda de estabilizacion lateral."],
["Suitcase Carry","Oblicuos","Core","Mancuernas","Antiflexion lateral dinamica","Intermedio","Una mancuerna/kettlebell a un lado, caminar recto."],
["Overhead Carry","Core","Hombros, Trapecios","Mancuernas","Estabilizacion con carga sobre cabeza","Intermedio","Mancuerna sobre la cabeza, caminar recto."],
["Cable Crunch","Recto Abdominal","Oblicuos","Cable","Flexion de columna","Principiante","De rodillas frente a la polea alta, flexionar el tronco."],
["Rollout con Barra","Core","Dorsal, Hombros","Barra","Antiextension dinamica","Avanzado","Barra en el suelo con discos, empujar hacia adelante."],
["Stir the Pot","Core","Hombros","Peso Corporal","Estabilizacion rotatoria","Avanzado","Antebrazos en pelota de estabilidad, mover en circulos pequeños."],

# ══════════════════════════════════════════════════════════════════════════════
# ANTEBRAZO / GRIP
# ══════════════════════════════════════════════════════════════════════════════
["Curl de Muneca con Barra Supino","Flexores del Antebrazo","Biceps","Barra","Flexion de muneca","Principiante","Antebrazos apoyados en banco, flexionar la muneca."],
["Curl de Muneca con Barra Prono","Extensores del Antebrazo","","Barra","Extension de muneca","Principiante","Antebrazos apoyados, extender la muneca."],
["Curl de Muneca con Mancuerna","Flexores del Antebrazo","","Mancuernas","Flexion de muneca","Principiante","Antebrazo apoyado en el muslo."],
["Extension de Muneca con Mancuerna","Extensores del Antebrazo","","Mancuernas","Extension de muneca","Principiante","Antebrazo apoyado, extender la muneca hacia arriba."],
["Curl de Muneca con Cable","Flexores del Antebrazo","","Cable","Flexion de muneca","Principiante","Polea baja, tension constante."],
["Extension de Muneca con Cable","Extensores del Antebrazo","","Cable","Extension de muneca","Principiante","Polea baja, extension de muneca."],
["Dead Hang (Barra)","Antebrazo","Dorsal, Hombros","Peso Corporal","Grip isometrico","Principiante","Colgar de la barra el mayor tiempo posible."],
["Dead Hang con Toalla","Antebrazo","Dorsal","Peso Corporal","Grip isometrico grueso","Intermedio","Toalla sobre la barra, agarrar la toalla."],
["Farmer Walk (Grip)","Antebrazo","Trapecios, Core","Mancuernas","Grip dinamico","Principiante","Mancuernas pesadas a los costados, caminar."],
["Plate Pinch Carry","Antebrazo","","Maquina","Grip de pellizco","Intermedio","Sujetar dos discos juntos con los dedos, caminar."],
["Reverse Curl con Barra","Braquiorradial","Extensores del Antebrazo","Barra","Flexion de codo prono","Principiante","Agarre prono, curvar la barra hacia arriba."],
["Towel Pull-up","Antebrazo","Dorsal, Biceps","Peso Corporal","Traccion vertical con grip grueso","Avanzado","Toallas sobre la barra, hacer dominadas sujetando las toallas."],
["Wrist Roller","Flexores y Extensores del Antebrazo","","Maquina","Flexion/extension alternada de muneca","Principiante","Rollo con cuerda y peso, enrollar y desenrollar."],

# ══════════════════════════════════════════════════════════════════════════════
# ADUCTORES
# ══════════════════════════════════════════════════════════════════════════════
["Aduccion de Cadera en Maquina","Aductores","Gluteos","Maquina","Aduccion de cadera","Principiante","Juntar las piernas contra la resistencia."],
["Sentadilla Sumo","Aductores","Gluteos, Cuadriceps","Barra","Sentadilla amplia","Intermedio","Pies muy abiertos, puntas hacia afuera."],
["Sentadilla Sumo con Mancuerna (Plie)","Aductores","Gluteos, Cuadriceps","Mancuernas","Sentadilla amplia","Principiante","Mancuerna vertical entre las piernas."],
["Aduccion de Cadera con Cable","Aductores","Gluteos","Cable","Aduccion de cadera","Principiante","Tobillera, llevar la pierna hacia la linea media."],
["Aduccion con Banda (Clamshell Inverso)","Aductores","","Bandas Elasticas","Aduccion de cadera","Principiante","Tumbado de lado, cerrar la rodilla contra la banda."],
["Zancada Lateral","Aductores","Gluteos, Cuadriceps","Peso Corporal","Zancada lateral","Principiante","Paso lateral amplio, doblar la rodilla de carga."],
["Lateral Squat","Aductores","Gluteos, Cuadriceps","Peso Corporal","Sentadilla lateral","Intermedio","Ampliar la zancada lateral y bajar profundo."],

# ══════════════════════════════════════════════════════════════════════════════
# ESPALDA BAJA / LUMBAR
# ══════════════════════════════════════════════════════════════════════════════
["Extension de Espalda Baja en Maquina","Espalda Baja","Gluteos, Isquiotibiales","Maquina","Extension de columna","Principiante","Sentado, extender la columna contra la resistencia."],
["Hiperextensiones en Banco Romano","Espalda Baja","Gluteos, Isquiotibiales","Maquina","Extension de cadera/columna","Principiante","Banco romano, bajar el torso y subir a la horizontal."],
["Hiperextensiones con Peso","Espalda Baja","Gluteos, Isquiotibiales","Maquina","Extension de cadera/columna","Intermedio","Sostener un plato o mancuerna frente al pecho."],
["Hiperextension Reversa","Gluteos","Isquiotibiales, Espalda Baja","Maquina","Extension de cadera invertida","Intermedio","Elevar las piernas en lugar del torso."],
["Superman","Espalda Baja","Gluteos, Isquiotibiales","Peso Corporal","Extension de columna y cadera","Principiante","Tumbado boca abajo, elevar brazos y piernas simultaneamente."],
["Superman con Brazos al Costado","Espalda Baja","Gluteos","Peso Corporal","Extension de columna","Principiante","Version mas facil con brazos pegados al cuerpo."],
["Puente de Gluteos Isometrico","Espalda Baja","Gluteos, Isquiotibiales","Peso Corporal","Isometria de cadera","Principiante","Puente de gluteos mantenido en la parte superior."],
["Deadbug (Estabilizacion Lumbar)","Espalda Baja","Core","Peso Corporal","Estabilizacion antiextension","Principiante","Tumbado, extender brazo y pierna opuestos manteniendo la zona lumbar."],
["Bird Dog","Espalda Baja","Gluteos, Core","Peso Corporal","Estabilizacion en cuadrupedia","Principiante","En cuatro apoyos, extender brazo y pierna opuestos."],
["Cat-Cow (Movilidad)","Espalda Baja","Core","Peso Corporal","Movilidad de columna","Principiante","En cuatro apoyos, alternar flexion y extension de columna."],

# ══════════════════════════════════════════════════════════════════════════════
# KETTLEBELL (especificos)
# ══════════════════════════════════════════════════════════════════════════════
["Swing con Kettlebell (Dos Manos)","Gluteos","Isquiotibiales, Core, Espalda Baja","Kettlebell","Bisagra de cadera explosiva","Intermedio","Bisagra explosiva, swing hasta la altura de los hombros."],
["Swing con Kettlebell (Un Brazo)","Gluteos","Isquiotibiales, Core, Hombros","Kettlebell","Bisagra de cadera explosiva unilateral","Intermedio","Mayor rotacion y estabilizacion."],
["Turkish Get-Up","Core","Hombros, Gluteos, Cuadriceps","Kettlebell","Patron completo de movimiento","Avanzado","Del suelo a de pie con kettlebell sobre la cabeza."],
["Goblet Squat con Kettlebell","Cuadriceps","Gluteos, Core","Kettlebell","Sentadilla frontal","Principiante","Kettlebell frente al pecho, torso erguido."],
["Press con Kettlebell (Un Brazo)","Hombros","Triceps, Core","Kettlebell","Empuje vertical unilateral","Intermedio","Press sobre la cabeza desde posicion de rack."],
["Remo con Kettlebell","Espalda Alta","Biceps","Kettlebell","Traccion horizontal","Principiante","Un brazo, igual que el remo con mancuerna."],
["Clean con Kettlebell","Gluteos","Core, Hombros, Isquiotibiales","Kettlebell","Movimiento olimpico","Intermedio","Desde el suelo hasta posicion de rack en un movimiento."],
["Snatch con Kettlebell","Gluteos","Core, Hombros, Isquiotibiales","Kettlebell","Movimiento olimpico","Avanzado","Desde el suelo hasta brazo extendido sobre la cabeza."],
["Windmill con Kettlebell","Oblicuos","Core, Hombros, Isquiotibiales","Kettlebell","Flexion lateral con carga sobre cabeza","Avanzado","Kettlebell sobre la cabeza, inclinar el torso lateralmente."],
["Floor Press con Kettlebell","Pecho","Triceps, Hombros","Kettlebell","Empuje horizontal en suelo","Principiante","Acostado en suelo, press con kettlebell."],
["Renegade Row","Espalda Alta","Core, Hombros","Kettlebell","Traccion horizontal en plancha","Avanzado","Plancha con manos en kettlebells, remar alternando."],
["Deadlift con Kettlebell","Espalda Baja","Gluteos, Isquiotibiales","Kettlebell","Bisagra de cadera","Principiante","Igual que el peso muerto pero con kettlebell entre los pies."],
["Around the World","Hombros","Core","Kettlebell","Circunduccion de hombro","Intermedio","Pasar la kettlebell alrededor del cuerpo."],
["Halo con Kettlebell","Hombros","Core, Cuello","Kettlebell","Circunduccion sobre la cabeza","Principiante","Pasar la kettlebell en circulos alrededor de la cabeza."],
["Sumo Deadlift con Kettlebell","Gluteos","Isquiotibiales, Aductores","Kettlebell","Bisagra de cadera amplia","Principiante","Postura sumo, kettlebell entre los pies."],

# ══════════════════════════════════════════════════════════════════════════════
# BANDAS ELASTICAS
# ══════════════════════════════════════════════════════════════════════════════
["Sentadilla con Banda","Cuadriceps","Gluteos","Bandas Elasticas","Sentadilla","Principiante","Banda sobre los hombros o bajo los pies."],
["Hip Thrust con Banda","Gluteos","Isquiotibiales","Bandas Elasticas","Extension de cadera","Principiante","Banda sobre la cadera."],
["Peso Muerto con Banda","Espalda Baja","Gluteos, Isquiotibiales","Bandas Elasticas","Bisagra de cadera","Principiante","Banda bajo los pies, agarrar al nivel de las caderas."],
["Curl de Biceps con Banda","Biceps","Braquial","Bandas Elasticas","Flexion de codo","Principiante","Banda bajo los pies, curl completo."],
["Press de Hombros con Banda","Hombros","Triceps","Bandas Elasticas","Empuje vertical","Principiante","Banda bajo los pies, press sobre la cabeza."],
["Press de Banca con Banda","Pecho","Triceps, Hombros","Bandas Elasticas","Empuje horizontal","Principiante","Banda anclada detras, press horizontal."],
["Face Pull con Banda","Hombros (posterior)","Trapecios","Bandas Elasticas","Traccion diagonal","Principiante","Banda anclada a altura del pecho."],
["Abduccion de Cadera con Banda","Gluteos (medio)","","Bandas Elasticas","Abduccion de cadera","Principiante","Banda en tobillos o rodillas."],
["Monster Walk","Gluteos (medio)","TFL","Bandas Elasticas","Abduccion dinamica","Principiante","Banda en rodillas, pasos laterales."],
["Jalon al Pecho con Banda","Dorsal","Biceps, Romboides","Bandas Elasticas","Traccion vertical","Principiante","Banda anclada arriba, jalon hacia la clavicula."],
["Rotacion Externa con Banda","Manguito Rotador","Hombros (posterior)","Bandas Elasticas","Rotacion externa de hombro","Principiante","Codo a 90, rotar externamente."],
["Pull-Apart con Banda","Hombros (posterior)","Romboides, Trapecios","Bandas Elasticas","Retraccion escapular","Principiante","Sostener la banda frente al pecho y separar los brazos."],
["Good Morning con Banda","Isquiotibiales","Gluteos, Espalda Baja","Bandas Elasticas","Bisagra de cadera","Principiante","Banda sobre los hombros anclada al suelo."],
["Pallof Press con Banda","Core","Oblicuos","Bandas Elasticas","Antirrotacion","Principiante","Banda anclada lateral, empujar al frente resistiendo la rotacion."],
["Aduccion de Cadera con Banda","Aductores","Gluteos","Bandas Elasticas","Aduccion de cadera","Principiante","Banda en tobillo, llevar la pierna hacia la linea media."],
["Clamshell con Banda","Gluteos (medio)","","Bandas Elasticas","Rotacion externa de cadera","Principiante","Tumbado de lado, abrir como una almeja."],
["Extensión de Triceps con Banda","Triceps","","Bandas Elasticas","Extension de codo","Principiante","Banda anclada arriba, pushdown."],

# ══════════════════════════════════════════════════════════════════════════════
# TRX / SUSPENSION
# ══════════════════════════════════════════════════════════════════════════════
["TRX Row (Remo Invertido)","Espalda Alta","Biceps, Romboides","TRX / Suspension","Traccion horizontal","Principiante","Cuerpo inclinado, tirar el pecho hacia los mangos."],
["TRX Row Unilateral","Espalda Alta","Biceps","TRX / Suspension","Traccion horizontal unilateral","Intermedio","Un brazo, mayor rango y estabilizacion."],
["TRX Push-Up","Pecho","Triceps, Core","TRX / Suspension","Empuje horizontal inestable","Intermedio","Mayor inestabilidad que las flexiones normales."],
["TRX Curl de Biceps","Biceps","Core","TRX / Suspension","Flexion de codo en suspension","Principiante","Cuerpo inclinado, curl hacia la frente."],
["TRX Extension de Triceps","Triceps","Core","TRX / Suspension","Extension de codo en suspension","Intermedio","Cuerpo inclinado, doblar codos hacia la frente."],
["TRX Pistol Squat Asistido","Cuadriceps","Gluteos","TRX / Suspension","Sentadilla unilateral asistida","Intermedio","Usar los mangos como soporte."],
["TRX Plank","Core","Hombros","TRX / Suspension","Isometria inestable","Intermedio","Pies en los mangos, cuerpo recto."],
["TRX Pike","Core","Hombros","TRX / Suspension","Flexion de cadera en suspension","Avanzado","Pies en los mangos, llevar caderas hacia arriba."],
["TRX Abduccion de Cadera","Gluteos (medio)","","TRX / Suspension","Abduccion de cadera en suspension","Intermedio","Un pie en el mango, elevar la pierna lateralmente."],
["TRX Curl de Isquiotibiales","Isquiotibiales","Gluteos","TRX / Suspension","Flexion de rodilla en suspension","Intermedio","Tumbado, pies en los mangos, elevar caderas y doblar rodillas."],
["TRX Face Pull","Hombros (posterior)","Trapecios","TRX / Suspension","Traccion diagonal en suspension","Principiante","Cuerpo inclinado, tirar los mangos hacia la cara."],
["TRX Squat Jump","Cuadriceps","Gluteos","TRX / Suspension","Sentadilla explosiva asistida","Intermedio","Usar los mangos para asistir el salto."],
["TRX Chest Fly","Pecho","Hombros (anterior)","TRX / Suspension","Aislamiento horizontal en suspension","Intermedio","Cuerpo inclinado, apertura en suspension."],
["TRX Mountain Climbers","Core","Hombros, Cuadriceps","TRX / Suspension","Dinamico de core en suspension","Intermedio","Pies en los mangos, llevar rodillas al pecho."],
["TRX Lunge","Cuadriceps","Gluteos","TRX / Suspension","Zancada en suspension","Intermedio","Un pie en el mango, zancada hacia adelante."],

# ══════════════════════════════════════════════════════════════════════════════
# MOVIMIENTOS OLIMPICOS / FUNCIONALES
# ══════════════════════════════════════════════════════════════════════════════
["Clean and Jerk","Gluteos","Cuadriceps, Hombros, Core, Isquiotibiales","Barra","Movimiento olimpico completo","Avanzado","Levantar la barra del suelo a los hombros (clean) y luego sobre la cabeza (jerk)."],
["Snatch","Gluteos","Cuadriceps, Hombros, Core, Isquiotibiales","Barra","Movimiento olimpico completo","Avanzado","Levantar la barra del suelo a sobre la cabeza en un solo movimiento."],
["Power Clean","Gluteos","Cuadriceps, Isquiotibiales, Core, Hombros","Barra","Movimiento de potencia","Avanzado","Clean sin caer bajo la barra, quedando de pie."],
["Power Snatch","Gluteos","Cuadriceps, Isquiotibiales, Hombros","Barra","Movimiento de potencia","Avanzado","Snatch sin caer en sentadilla."],
["Hang Clean","Gluteos","Cuadriceps, Isquiotibiales, Hombros","Barra","Movimiento olimpico desde hang","Avanzado","Clean comenzando desde la posicion de hang (barra a la altura del muslo)."],
["Hang Power Clean","Gluteos","Cuadriceps, Isquiotibiales, Hombros","Barra","Movimiento de potencia desde hang","Avanzado","Power clean desde posicion de hang."],
["High Pull","Trapecios","Hombros, Isquiotibiales, Gluteos","Barra","Tiracion explosiva vertical","Avanzado","Tiracion de barra desde el suelo hasta la barbilla con impulso de caderas."],
["Push Press con Barra","Hombros","Triceps, Cuadriceps, Core","Barra","Empuje vertical con impulso","Intermedio","Ligero impulso de piernas para dar inercia a la barra."],
["Push Jerk","Hombros","Triceps, Cuadriceps, Core","Barra","Empuje vertical olimpico","Avanzado","Impulso de piernas, caer bajo la barra con brazos extendidos."],
["Split Jerk","Hombros","Cuadriceps, Core","Barra","Empuje vertical con separacion de piernas","Avanzado","Caer con una pierna adelante y otra atras al extender los brazos."],
["Landmine Squat to Press","Cuadriceps","Gluteos, Hombros, Core","Barra","Compuesto de sentadilla + press","Intermedio","Un extremo de la barra anclado, sentadilla y press en un movimiento."],

# ══════════════════════════════════════════════════════════════════════════════
# MAQUINAS POCO CONOCIDAS / ESPECIFICAS
# ══════════════════════════════════════════════════════════════════════════════
["Pendulum Squat","Cuadriceps","Gluteos","Maquina","Sentadilla en arco pendular","Intermedio","Maquina que sigue un arco, excelente para cuadriceps sin carga en columna."],
["Belt Squat","Cuadriceps","Gluteos","Maquina","Sentadilla sin carga en columna","Intermedio","La carga se cuelga del cinto, sin carga axial."],
["Reverse Hyper (Extension Inversa)","Gluteos","Isquiotibiales, Espalda Baja","Maquina","Extension de cadera invertida","Intermedio","Tumbado boca abajo, elevar las piernas cargadas."],
["Maquina de Aduccion (Inner Thigh)","Aductores","Gluteos","Maquina","Aduccion de cadera","Principiante","Juntar las piernas contra la resistencia."],
["Maquina de Abduccion (Outer Thigh)","Gluteos (medio)","TFL","Maquina","Abduccion de cadera","Principiante","Separar las piernas contra la resistencia."],
["Glute-Ham Raise (GHR)","Isquiotibiales","Gluteos, Espalda Baja","Maquina","Flexion de rodilla + extension de cadera","Avanzado","En GHD, bajar y subir el torso activando ambos isquio y gluteo."],
["Sissy Squat","Cuadriceps","","Maquina","Extension de rodilla aislada","Avanzado","Rodillas avanzan adelante mientras el torso cae hacia atras."],
["Maquina de Biceps Femoral (Lying)","Isquiotibiales","Gastrocnemio","Maquina","Flexion de rodilla guiada","Principiante","Version horizontal, muy comun en gimnasios."],
["Maquina de Biceps Femoral (Seated)","Isquiotibiales","","Maquina","Flexion de rodilla sentado","Principiante","Version sentada, diferente curva de tension."],
["Maquina de Extension de Cuadriceps","Cuadriceps","","Maquina","Extension de rodilla","Principiante","Classic leg extension machine."],
["Prensa Vertical (Vertical Leg Press)","Cuadriceps","Gluteos","Maquina","Empuje vertical","Principiante","Pies hacia arriba, empujar la plataforma verticalmente."],
["Maquina de Pantorrillas Sentado (Soleo)","Soleo","Gastrocnemio","Maquina","Flexion plantar sentado","Principiante","Rodillas a 90, mayor enfasis en soleo."],
["Maquina de Pantorrillas De Pie","Gastrocnemio","Soleo","Maquina","Flexion plantar de pie","Principiante","Hombros bajo los pads."],
["Maquina de Triceps Dip Asistida","Triceps","Pecho, Hombros","Maquina","Empuje vertical asistido","Principiante","Contrarresta parte del peso corporal para facilitar los fondos."],
["Maquina de Remo Independiente Nautilus","Espalda Alta","Biceps, Romboides","Maquina","Traccion horizontal articulada","Intermedio","Palancas independientes, patron muy natural."],
["Pec Deck Invertido (Reverse Pec Deck)","Hombros (posterior)","Romboides","Maquina","Aislamiento posterior","Principiante","Version invertida de la maquina de aperturas."],
["Cable Crossover Machine","Pecho","Hombros (anterior)","Cable","Cruce de poleas","Principiante","Maquina de poleas ajustables a distintas alturas."],
["Maquina de Biceps Curl (Preacher)","Biceps","Braquial","Maquina","Flexion de codo apoyada","Principiante","Brazo apoyado en pad inclinado, movimiento guiado."],
["Maquina de Hombros (Lateral Raise)","Hombros (lateral)","Trapecios","Maquina","Aislamiento lateral","Principiante","Tension constante a lo largo de todo el rango."],
["Maquina de Rotacion de Tronco","Oblicuos","Core","Maquina","Rotacion de tronco guiada","Principiante","Sentado, rotar el tronco contra la resistencia."],
["Maquina Smith (Usos Generales)","Variable","Variable","Maquina Smith","Multiple","Principiante","Barra guiada en carril, util para aprender patrones."],
["Maquina de Back Extension","Espalda Baja","Gluteos, Isquiotibiales","Maquina","Extension de columna","Principiante","Version sentada de la extension de columna."],
["Jammer (Press Explosivo en Maquina)","Hombros","Pecho, Triceps, Core","Maquina","Empuje explosivo diagonal","Avanzado","Maquina de prensa explosiva en angulo diagonal."],
["Rower (Maquina de Remo)","Espalda Alta","Biceps, Cuadriceps, Gluteos","Maquina","Cardio-fuerza de remo","Intermedio","Maquina de remo de resistencia, trabaja cuerpo completo."],
["Maquina de Esqui (SkiErg)","Core","Hombros, Espalda Alta, Cuadriceps","Maquina","Cardio-fuerza overhead","Intermedio","Doble movimiento de jalado hacia abajo, imita el esqui nordico."],
["Maquina de Escalada (VersaClimber)","Core","Hombros, Cuadriceps, Gluteos","Maquina","Cardio-fuerza de escalada","Intermedio","Movimiento alternado de brazos y piernas en vertical."],
["Maquina Battle Rope","Core","Hombros, Espalda Alta","Maquina","Cardio-fuerza con cuerdas","Intermedio","Cuerdas pesadas, movimientos ondulatorios de brazos."],
["AirBike (Bicicleta de Aire)","Cuadriceps","Gluteos, Hombros, Core","Maquina","Cardio-fuerza cuerpo completo","Principiante","Bicicleta con palancas para los brazos, resistencia por aire."],
]

HEADER_BG = "1E293B"
HEADER_FG = "FFFFFF"
EVEN_ROW_BG = "F8FAFC"
ODD_ROW_BG = "FFFFFF"

EQUIP_COLORS = {
    "Barra": "FEF3C7",
    "Mancuernas": "DBEAFE",
    "Maquina": "F3E8FF",
    "Maquina Smith": "EDE9FE",
    "Barra Hexagonal": "FDE68A",
    "Cable": "DCFCE7",
    "Kettlebell": "FFE4E6",
    "Bandas Elasticas": "FEF9C3",
    "TRX / Suspension": "E0F2FE",
    "Peso Corporal": "F1F5F9",
}

DIFF_COLORS = {
    "Principiante": "DCFCE7",
    "Principiante": "DCFCE7",
    "Intermedio": "FEF9C3",
    "Avanzado": "FFE4E6",
}

thin = Side(style='thin', color='E2E8F0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(name='Arial', bold=True, color=HEADER_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 32

for i, ex in enumerate(exercises, 2):
    row_data = [i - 1] + ex
    equip = ex[3]
    diff = ex[5]
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=(col == 8))
        if col == 5:
            color = EQUIP_COLORS.get(equip, "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif col == 7:
            color = DIFF_COLORS.get(diff, "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif i % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=EVEN_ROW_BG)
        else:
            cell.fill = PatternFill("solid", fgColor=ODD_ROW_BG)
    ws.row_dimensions[i].height = 40

widths = [5, 40, 22, 36, 18, 26, 14, 52]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
ws.freeze_panes = "A2"

# Leyenda
ws2 = wb.create_sheet("Leyenda")
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
for col in range(1, 3):
    c = ws2.cell(1, col, ["EQUIPAMIENTO", "Color"][col-1])
    c.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
for r, (eq, color) in enumerate(EQUIP_COLORS.items(), 2):
    ws2.cell(r, 1, eq).font = Font(name='Arial', size=9)
    ws2.cell(r, 2, "").fill = PatternFill("solid", fgColor=color)
    ws2.row_dimensions[r].height = 20
r_offset = len(EQUIP_COLORS) + 3
for col in range(1, 3):
    c = ws2.cell(r_offset, col, ["DIFICULTAD", "Color"][col-1])
    c.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
for r, (diff, color) in enumerate({"Principiante":"DCFCE7","Intermedio":"FEF9C3","Avanzado":"FFE4E6"}.items(), r_offset+1):
    ws2.cell(r, 1, diff).font = Font(name='Arial', size=9)
    ws2.cell(r, 2, "").fill = PatternFill("solid", fgColor=color)
    ws2.row_dimensions[r].height = 20

output_path = r"C:\Users\felds\Desktop\heman\ejercicios_heman.xlsx"
wb.save(output_path)
print(f"Guardado: {output_path} — {len(exercises)} ejercicios")
